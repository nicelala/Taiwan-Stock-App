import io

import pytest
from openpyxl import load_workbook


def test_export_dividends_csv_returns_bom_and_chinese_header(client, db_session):
    """
    dividends CSV export:
    - status_code 200
    - Content-Disposition exists and ends with .csv
    - response starts with UTF-8 BOM
    - first line contains Chinese headers
    """
    # 允許空資料也能匯出：至少要有 header + BOM
    resp = client.get("/api/v1/dividends/export?format=csv")

    assert resp.status_code == 200
    cd = resp.headers.get("content-disposition", "")
    assert "attachment" in cd.lower()
    assert cd.lower().endswith('.csv"') or ".csv" in cd.lower()

    # BOM check (EF BB BF)
    assert resp.content[:3] == b"\xef\xbb\xbf"

    # header line should be Chinese
    text = resp.content.decode("utf-8", errors="replace")
    first_line = text.splitlines()[0]
    assert "股票代號" in first_line
    assert "公司名稱" in first_line
    assert "總股利" in first_line


def test_export_dividends_xlsx_returns_workbook_and_chinese_header(client, db_session):
    """
    dividends XLSX export:
    - status_code 200
    - Content-Disposition exists and ends with .xlsx
    - Content-Type is xlsx
    - first row contains Chinese headers
    """
    resp = client.get("/api/v1/dividends/export?format=xlsx")

    assert resp.status_code == 200
    cd = resp.headers.get("content-disposition", "")
    assert "attachment" in cd.lower()
    assert cd.lower().endswith('.xlsx"') or ".xlsx" in cd.lower()

    ctype = resp.headers.get("content-type", "").lower()
    assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in ctype

    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    # Chinese headers
    assert "股票代號" in headers
    assert "公司名稱" in headers
    assert "總股利" in headers


def test_export_refresh_logs_csv_returns_bom_and_api_key_header(client, db_session):
    """
    refresh logs CSV export:
    - status_code 200
    - Content-Disposition exists and ends with .csv
    - response starts with UTF-8 BOM
    - first line is API key header
    """
    resp = client.get("/api/v1/admin/refresh/logs/export?format=csv")

    assert resp.status_code == 200
    cd = resp.headers.get("content-disposition", "")
    assert "attachment" in cd.lower()
    assert cd.lower().endswith('.csv"') or ".csv" in cd.lower()

    assert resp.content[:3] == b"\xef\xbb\xbf"

    text = resp.content.decode("utf-8", errors="replace")
    first_line = text.splitlines()[0]

    # header uses API keys (not Chinese)
    assert "id" in first_line
    assert "job_name" in first_line
    assert "market" in first_line
    assert "status" in first_line
    assert "trigger_source" in first_line


def test_export_dividends_csv_too_large_returns_422(monkeypatch, client, db_session):
    """
    dividends CSV export should enforce max 200000.
    Use monkeypatch to avoid heavy DB seeding.
    """
    # IMPORTANT: adjust import path if your module path differs
    from app.repositories.dividend_repository import DividendRepository

    original_search = DividendRepository.search

    def fake_search(self, *args, **kwargs):
        # first query uses limit=1; we return a huge total_count to trigger limit check
        return [], 200_001

    monkeypatch.setattr(DividendRepository, "search", fake_search)

    resp = client.get("/api/v1/dividends/export?format=csv")
    assert resp.status_code == 422

    data = resp.json()
    assert data["detail"]["code"] == "EXPORT_TOO_LARGE"

    monkeypatch.setattr(DividendRepository, "search", original_search)


def test_export_dividends_xlsx_too_large_returns_422(monkeypatch, client, db_session):
    """
    dividends XLSX export should enforce max 20000.
    Use monkeypatch to avoid heavy DB seeding.
    """
    from app.repositories.dividend_repository import DividendRepository

    original_search = DividendRepository.search

    def fake_search(self, *args, **kwargs):
        return [], 20_001

    monkeypatch.setattr(DividendRepository, "search", fake_search)

    resp = client.get("/api/v1/dividends/export?format=xlsx")
    assert resp.status_code == 422

    data = resp.json()
    assert data["detail"]["code"] == "EXPORT_TOO_LARGE"

    monkeypatch.setattr(DividendRepository, "search", original_search)